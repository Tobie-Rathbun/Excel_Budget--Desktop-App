#created by Tobie Rathbun

from openpyxl import Workbook
from openpyxl.styles import PatternFill
from openpyxl.styles.differential import DifferentialStyle
from openpyxl.styles import Font, Color


#creates workbook
wb = Workbook()


#creates worksheet and names it
ws = wb.active
ws.title = "Budget Calculator"


black = Color(indexed=0)
white = Color(indexed=1)
red = Color(indexed=2)
green = Color(indexed=3)
yellow = Color(indexed=5)
light = Color(indexed=22)
dark = Color(indexed=23)

color_sel = black



#creates this far down on excel
#ws['E20'].fill = PatternFill(start_color=red, fill_type='solid')



#black header
def blockSel():
	for row in ws.iter_rows(
		max_row = row_max, 
		min_row = row_min,
		max_col = 5):
		for cell in row:
			cell.fill = PatternFill(
				start_color = color_sel, 
				fill_type = 'solid',
				)

			
ws.merge_cells('A1:E1')




row_min = 1
row_max = 1

blockSel()

color_sel = light
row_min = 2
row_max = 5

blockSel()



a1 = ws['A1']
a1.value = 'Income'

b2 = ws['B2']
b2.value = 'Hourly Wage'

header_text = a1
header_text.font = Font(
	color = white, 
	bold = True,)
	
body_text = b2
body_text.font = Font(
	color = black, 
	bold = False,)








#saves document
wb.save('budget.xlsx')





